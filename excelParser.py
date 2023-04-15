from openpyxl import load_workbook


def parse(name: str):
    wb = load_workbook(name)
    ws = wb.active

    courses_dict = dict()

    for col in ws.iter_cols():
        print(col[0].value)

    for row in ws.iter_rows():
        class_info = []
        for cell in row:
            class_info.append(cell.value)
        print(class_info)
